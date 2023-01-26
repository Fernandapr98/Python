# My first Excel
import random
import openpyxl
import pycountry

participants = []
amount = []
country = []

with open ("Compañeros.txt", "r") as x:
    for line in x:
        participants.append(line.strip())
money = random.randint(1, 100000)
for i in range(len(participants)):
    random_money = random.randint(1, 100000)
    money += random_money
    amount.append(random_money)
for i in range(len(participants)):
    random_country = random.randint(0, 244)
    country.append(list(pycountry.countries)[random_country].name)
    
wb = openpyxl.Workbook()
ws = wb.active
#Títulos de las columnas
ws['A1'] = "Participant"
ws['B1'] = "Amount"
ws['C1'] = "Country"

for i, j, k in zip(participants, amount, country):
    ws.cell(row=participants.index(i)+2, column = 1).value = i
    ws.cell(row=participants.index(i)+2, column = 2).value = j
    ws.cell(row=participants.index(i)+2, column = 3).value = k
# Save the file
ws.cell(row=len(participants)+2, column = 2).value = money
wb.save("MyFirstExcel.xlsx")