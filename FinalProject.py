# FINAL PROJECT

# IMPORTS
import pip_system_certs.wrapt_requests
import requests
import json
import openpyxl
from whoisapi import *

# ARCHIVO EXCEL
wb = openpyxl.Workbook()
ws = wb.active 

# Titles in Excel:
ws['A1'] = "path filename"
ws['B1'] = "domainName"
ws['C1'] = "registrarName"
ws['D1'] = "contactEmail"
ws['E1'] = "registryData.createdDate"
ws['F1'] = "registrant.country"

# VARIABLES
api_key_fer = "at_hMvDm8xn4Sjhxc9TYb7REg2scXR7p"
api_endpoint = "https://www.whoisxmlapi.com/whoisserver/WhoisService?apiKey=at_hMvDm8xn4Sjhxc9TYb7REg2scXR7p&domainName=google.com"


# API Request
def request(ip_address, api_key_fer):
    try:
        client = Client(api_key= api_key_fer)
        client.parameters.output_format = 'json'
        # Get raw API response
        response = client.raw_data(api_endpoint)
        json_file = json.loads(response)
        print(json_file['WhoisRecord']['domainName'])
        createFile("Desktop/"+str(json_file['WhoisRecord']['domainName']+".json", response))
    
    except Exception as e:
        print("Sorry :( an error has ocurred during your request for the IP address "+ip_address+ str(e)+"\n")
        createFileError("FailedIPs.txt",str(json_file['WhoisRecord']['domainName']))
        
    updateExcel(json_file)

path_filename = ""
domainName = ""
registrarName = ""
contactEmail = " "
registryData_createdDate = " "
registrant_country = " "
    
def updateExcel(json_file):
    path_filename = ""
    domainName = ""
    registrarName = ""
    contactEmail = " "
    registryData_createdDate = " "
    registrant_country = " "
    
    # Variables with Value 
    path_filename = request(1, json_file, "Desktop/"+json_file['WhoisRecord']['domainName'])
    domainName = request(2, json_file, path_filename)
    registrarName = request(3, json_file, path_filename)
    contactEmail = request(4, json_file, path_filename)
    registryData_createdDate = request(5, json_file, path_filename)
    registrant_country = request(6, json_file, path_filename)

ws.append([path_filename, domainName, registrarName, contactEmail, registryData_createdDate, registrant_country])
wb.save("IPsResponse.xlsx") 
   

# Cheking the IPs
def test(x, json_file, path):
    if x == 1:
        try:
            return "Desktop/"+str(json_file["WhoisRecord"]["domainName"])
        except Exception as e:
            print("Error in"+path+" "+str(e)+" ")
    if x == 2:
        try:
            return str(json_file["WhoisRecord"]["domainName"])
        except Exception as e:
            print("Error in"+path+" "+str(e)+" ")
    if x == 3:
        try:
            return str(json_file["WhoisRecord"]["registrarName"])
        except Exception as e:
            print("Error in"+path+" "+str(e)+" ")
    if x == 4:
        try:
            return "Desktop/"+str(json_file["WhoisRecord"]["contactEmail"])
        except Exception as e:
            print("Error in"+path+" "+str(e)+" ")
    if x == 5:
        try:
            return "Desktop/"+str(json_file["WhoisRecord"]["registryData"]["createdDate"])
        except Exception as e:
            print("Error in"+path+" "+str(e)+" ")
    if x == 6:
        try:
            return "Desktop/"+str(json_file["WhoisRecord"]["registryData"]["registrant"]["country"])
        except Exception as e:
            print("Error in"+path+" "+str(e)+" ")
    else:
        return ""
        
# Keep all the IPs in a file
ip_list = []
with open ("lista100IPs.txt", "r") as x:
    for line in x:
        ip_list.append(line.strip())

def createFile(name, value):
    f = open(name, "w")
    json.dump(value, f, indent=4)
    f.close()
    
def createFileError(name, value):
    f = open(name, "w")
    f.write(value)
    f.close()










